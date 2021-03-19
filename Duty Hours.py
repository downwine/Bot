import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import re

path_to_file = 'dezhurstvo4etazh.xlsx'

current_datetime = datetime.datetime.now()
wb = openpyxl.load_workbook(path_to_file)  # Грузим наш график
sheets_list = wb.sheetnames  # Получаем список всех листов в файле
sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
a = sheet_active.cell(row=current_datetime.day+4, column=1).value
print("Дежурит:", a)


path_to_file = 'prozhivayuschie.xlsx'

search_text = a
print('Ищем:', search_text)

wb = openpyxl.load_workbook(path_to_file)  # Грузим наш список
sheets_list = wb.sheetnames  # Получаем список всех листов в файле
sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
row_max = sheet_active.max_row  # Получаем количество столбцов
# print(type(row_max))
column_max = sheet_active.max_column  # Получаем количество строк


row_min = 1  # Переменная, отвечающая за номер строки
column_min = 1  # Переменная, отвечающая за номер столбца


row_min_min = row_min
row_max_max = row_max
while row_min_min <= row_max_max:
    row_min_min = str(row_min_min)

    word_column = get_column_letter(column_min)
    word_column = str(word_column)
    word_cell = word_column + row_min_min

    data_from_cell = sheet_active[word_cell].value
    data_from_cell = str(data_from_cell)
    # print(data_from_cell)
    regular = search_text
    result = re.findall(regular, data_from_cell)
    if len(result) > 0:
        b = sheet_active.cell(row=int(word_cell[1:]), column=2).value
        print("id:", b)
    row_min_min = int(row_min_min)
    row_min_min = row_min_min + 1


if (current_datetime.hour==0 and current_datetime.minute==1):
    print("id:", b)