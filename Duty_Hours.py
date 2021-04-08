def search_in_table(row_max, searching_element, sheet_active, column):
    """
    Функция для поиска данных в указанном столбце
    :param row_max: количество строк в таблице
    :param searching_element: элемент, поиск которого выполняется
    :param sheet_active: лист с которым мы работаем в данный момент
    :param column: столбец, в котором выполняется поиск
    :return: список всех найденных элементов
    """
    from openpyxl.utils import get_column_letter
    import re
    row_min = 1  # Переменная, отвечающая за номер строки
    column = column  # Переменная, отвечающая за номер столбца
    found_texts = []  # Создаем пустой список для записи найденных элементов

    searching_str = str(searching_element)
    row_min_min = row_min
    row_max_max = row_max
    while row_min_min <= row_max_max:
        row_min_min = str(row_min_min)

        word_column = get_column_letter(column)
        word_column = str(word_column)
        word_cell = word_column + row_min_min

        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)
        # print(data_from_cell)
        regular = searching_str
        result = re.findall(regular, data_from_cell)
        row_min_min = int(row_min_min)
        row_min_min = row_min_min + 1
        if len(result) > 0:
            found_texts.append(word_cell)
    return found_texts


def search_id(searching_id):
    """
    Функция для поиска id в таблице
    :param searching_id: id, который необходимо найти
    :return: ФИО, соответствующее данному id
    """
    import openpyxl

    path_to_file = 'Список проживающих.xlsx'
    our_table = openpyxl.load_workbook(path_to_file)  # Грузим наш список
    sheet_active = our_table.active  # Начинаем работать с файлом
    row_max = sheet_active.max_row  # Получаем количество строк
    # print(type(row_max))
    word_cell = search_in_table(row_max, searching_id, sheet_active, 2)

    if not word_cell:
        return None
    else:
        sought_name = sheet_active.cell(row=int(word_cell[0][1:]), column=1).value
        return sought_name


# print(search_id(207826855))

def search_name(searching_name):
    """
    Функция для поиска ФИО в таблице
    :param searching_name: ФИО, которые необходимо найти
    :return: id, соответствующее данному ФИО
    """
    import openpyxl

    path_to_file = 'Список проживающих.xlsx'
    our_table = openpyxl.load_workbook(path_to_file)  # Грузим наш список
    sheet_active = our_table.active  # Начинаем работать с файлом
    row_max = sheet_active.max_row  # Получаем количество строк
    word_cell = search_in_table(row_max, searching_name, sheet_active, 1)

    if not word_cell:
        return None
    else:
        sought_id = sheet_active.cell(row=int(word_cell[0][1:]), column=2).value
        return sought_id


def duty_hours_today(flags):
    """    Функция, уведомляющая проживающих, что они дежурят сегодня    """
    import datetime
    import openpyxl

    current_datetime = datetime.datetime.now()
    if current_datetime.hour == 12 and current_datetime.minute == 0:  # Условие отправки сообщения 12:00
        ids_of_cleaners = []
        for i in 2, 3, 4, 5, 6, 7, 9:
            path_to_file = f'График дежурств {i} этаж.xlsx'
            try:
                our_table = openpyxl.load_workbook(path_to_file)  # Грузим наш график
            except:
                print(f'График дежурств {i} этаж.xlsx отсутствует')
                continue
            sheet_active = our_table.active  # Начинаем работать с файлом
            search_text = sheet_active.cell(row=current_datetime.day + 3, column=1).value

            path_to_file = 'Список проживающих.xlsx'
            our_table = openpyxl.load_workbook(path_to_file)  # Грузим наш список
            sheet_active = our_table.active  # Начинаем работать с файлом
            row_max = sheet_active.max_row  # Получаем количество строк
            word_cell = search_in_table(row_max, search_text, sheet_active, 1)
            # print(word_cell)
            if not word_cell:
                ids_of_cleaners.append(None)
            else:
                ids_of_cleaners.append(sheet_active.cell(row=int(word_cell[0][1:]), column=2).value)

        notify_cleaners(ids_of_cleaners, flags)


def notify_cleaners(ids_of_cleaners, flags):
    """
    Функция, отправляющая сообщения тем, кто дежурит в указанный день
    :param ids_of_cleaners: id проживающего
    :param flags
    """
    import vk_api
    from our_token import token
    from filling_docs import send_msg_without_keyboard

    i = 0
    vk_session = vk_api.VkApi(token=token)
    for idd in ids_of_cleaners:
        # Проверка комендант ли пользователь
        if vk_session.method("messages.isMessagesFromGroupAllowed",
                             {"group_id": 202823499, "user_id": idd})["is_allowed"]:
            if not flags[i]:
                send_msg_without_keyboard(idd, "Ты сегодня дежурный! Не забудь прибраться ;)")
                flags[i] = True
                i += 1


# flags = [False, False, False, False, False, False, False]
# print(duty_hours_today(flags))

def present_month():
    """
    Функция для получения текущего месяца на кириллице
    :return: месяц строкой на кириллице
    """
    import datetime

    current_datetime = datetime.datetime.now()
    month = current_datetime.month
    dict_of_months = {1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня', 7: 'июля',
                      8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'}
    return dict_of_months[month]


# print(present_month())

def duty_hours_when(received_id):
    """
    Функция для получениия даты дежурства проживающего
    :param received_id: id проживающего
    :return: список дат дежурства
    """
    import openpyxl
    word_cell = []
    sought_name = search_id(received_id)
    if not sought_name:
        return None
    surname = sought_name.split(" ")
    received_id = surname[0] + " " + surname[1][:1]
    for i in 2, 3, 4, 5, 6, 7, 9:
        path_to_file = f'График дежурств {i} этаж.xlsx'
        try:
            our_table = openpyxl.load_workbook(path_to_file)  # Грузим наш график
        except:
            print(f'График дежурств {i} этаж.xlsx отсутствует')
            continue
        sheet_active = our_table.active  # Начинаем работать с файлом
        row_max = sheet_active.max_row  # Получаем количество строк
        word_cell_test = search_in_table(row_max, received_id, sheet_active, 1)
        if not word_cell_test:
            received_id = surname[0]
            word_cell_test = search_in_table(row_max, received_id, sheet_active, 1)
        if word_cell_test:
            word_cell = word_cell_test
    cleaning_days = []
    if not word_cell:
        return None
    else:
        for i in range(len(word_cell)):
            cleaning_days.append(int(word_cell[i][1:]) - 3)
        return cleaning_days


# print(duty_hours_when(114075926))


# Перед вызовом функции необходимо закрыть файл
def delete_row(name):
    """
    Фунция для удаления строки в таблице(ФИО и id)
    :param name: ФИО проживающего, которого необходимо удолить
    """

    import openpyxl

    path_to_file = 'Список проживающих.xlsx'
    our_table = openpyxl.load_workbook(path_to_file)  # Грузим наш список
    sheet_active = our_table.active  # Начинаем работать с файлом
    row_max = sheet_active.max_row  # Получаем количество строк
    word_cells = search_in_table(row_max, name, sheet_active, 1)
    if not word_cells:
        return None
    else:
        word_cell = word_cells[0][1:]
        sheet_active.delete_rows(int(word_cell))
        our_table.save('Список проживающих.xlsx')


# delete_row('Михайлов Артем Алексеевич')

# Перед вызовом функции необходимо закрыть файл
def add_row(name, id_number):
    """
    Фунция для добавления строки в таблицу(ФИО и id)
    :param name: ФИО проживающего
    :param id_number: id проживающего
    """
    import openpyxl

    path_to_file = 'Список проживающих.xlsx'
    our_table = openpyxl.load_workbook(path_to_file)  # Грузим наш список
    sheet_active = our_table.active  # Начинаем работать с файлом
    row_max = sheet_active.max_row  # Получаем количество строк
    first_cell = "A" + str(row_max + 1)  # Ячейка с именем
    second_cell = "B" + str(row_max + 1)  # Ячейка с id
    sheet_active[first_cell].value = name
    sheet_active[second_cell].value = id_number
    our_table.save('Список проживающих.xlsx')
